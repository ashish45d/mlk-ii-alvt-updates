"""
MLL Parser Module
Business logic for parsing Microlok II .MLL listing files
"""

import re
from pathlib import Path


class MLLParser:
    """Parser for Microlok II .MLL listing files"""
    
    HEADERS = [
        'FILE', 'CRC', 'CHECKSUM', 'BOOLEAN_BITS_COUNT', 'NUMERIC_BITS_COUNT',
        'NUMERIC_BLOCKS_COUNT', 'ASSIGN_COUNT', 'NV_ASSIGN_COUNT', 'ASSIGN_TOTAL',
        'TIMER_BITS_COUNT', 'VITAL_INPUT_BOARDS', 'VITAL_OUTPUT_BOARDS',
        'NV_IN32_OUT32_BOARDS', 'MII_LINKS_COUNT', 'MII_NV_LINKS_COUNT',
        'COMM_LINKS_TOTAL', 'SYNC_MESSAGE_SIZE', 'VALIDATION_NOTES'
    ]
    
    @classmethod
    def validate_boolean_bits(cls, row):
        """Validate BOOLEAN_BITS_COUNT and return any warnings/errors"""
        try:
            boolean_bits = int(row.get('BOOLEAN_BITS_COUNT', 0))
            vi = int(row.get('VITAL_INPUT_BOARDS', 0))
            vo = int(row.get('VITAL_OUTPUT_BOARDS', 0))
            nv = int(row.get('NV_IN32_OUT32_BOARDS', 0))
            total_boards = vi + vo + nv
            
            messages = []
            
            if total_boards > 12:
                messages.append(f"WARNING: >12 boards ({total_boards} total)")
            
            # Calculate max allowed boolean bits based on board count
            max_bits = {
                2: 3600, 3: 3500, 4: 3400, 5: 3300, 6: 3200,
                7: 3100, 8: 3000, 9: 2900, 10: 2800, 11: 2700, 12: 2600
            }
            
            if total_boards > 12:
                max_allowed = 2600
            elif total_boards < 2:
                max_allowed = 3600
            else:
                max_allowed = max_bits.get(total_boards, 3600)
            
            if boolean_bits > max_allowed:
                messages.append(f"ERROR: Boolean bits ({boolean_bits}) > {max_allowed} allowed for {total_boards} boards")
                
            return messages
        except (ValueError, TypeError):
            return []
    
    @classmethod
    def validate_comm_links(cls, row):
        """Validate COMM_LINKS_TOTAL and return any warnings/errors"""
        try:
            comm_links = int(row.get('COMM_LINKS_TOTAL', 0))
            if comm_links > 32:
                return [f"ERROR: Communication links ({comm_links}) > 32"]
            return []
        except (ValueError, TypeError):
            return []
    
    @classmethod
    def validate_row(cls, row):
        """Validate a row of data and return a new row with validation notes"""
        # Make a copy to avoid modifying the original
        row = dict(row)
        
        # Run all validations
        validations = []
        validations.extend(cls.validate_boolean_bits(row))
        validations.extend(cls.validate_comm_links(row))
        
        # Add validation notes to the row
        row['VALIDATION_NOTES'] = '; '.join(validations) if validations else ''
        
        # Add warnings/errors to the relevant cells
        if 'BOOLEAN_BITS_COUNT' in row and any('Boolean bits' in v for v in validations):
            row['BOOLEAN_BITS_COUNT'] = f"{row['BOOLEAN_BITS_COUNT']} (!)"
            
        if 'COMM_LINKS_TOTAL' in row and any('Communication links' in v for v in validations):
            row['COMM_LINKS_TOTAL'] = f"{row['COMM_LINKS_TOTAL']} (!)"
        
        return row
    
    @staticmethod
    def read_mll_text(file_path):
        """Read listing file as text, robustly"""
        p = Path(file_path)
        for enc in ("utf-8", "latin-1"):
            try:
                return p.read_text(encoding=enc)
            except UnicodeDecodeError:
                continue
        return p.read_text(encoding="utf-8", errors="ignore")
    
    @staticmethod
    def extract_crc_checksum(text):
        """Return (CRC, CHECKSUM) from the first header line that contains both"""
        crc_regex = re.compile(r'CRC\s*=\s*([0-9A-Fa-f]+)')
        checksum_regex = re.compile(r'Checksum\s*=\s*([0-9A-Fa-f]+)')
        
        m_crc = crc_regex.search(text)
        m_ck = checksum_regex.search(text)
        return (m_crc.group(1) if m_crc else None,
                m_ck.group(1) if m_ck else None)
    
    @staticmethod
    def extract_boolean_bits_count(text):
        """BOOLEAN_BITS_COUNT: last index in Bit Usage Summary BEFORE Numeric Usage Summary"""
        boolean_usage_hdr = re.compile(r'(?is)\*{3}\s*Bit\s+Usage\s+Summary')
        numeric_usage_hdr = re.compile(r'(?is)\*{3}\s*Numeric\s+Usage\s+Summary')
        bit_use_row_tail = r'(?:NASGN|SYSTEM|CONFIG)\s+(?:VITAL|NON)\s+(?:INT|OUT)\s*$'
        bit_use_row = re.compile(rf'(?m)^\s*(\d+)\s+\S.*?\s+{bit_use_row_tail}')
        
        # 1) Explicit Boolean Usage Summary block
        m_bool = boolean_usage_hdr.search(text)
        if m_bool:
            m_block = re.search(r'(?is)\*{3}\s*Bit\s+Usage\s+Summary.*?(?=\n\s*\*{3}|\Z)', text)
            if m_block:
                block = m_block.group(0)
                idxs = [int(x) for x in re.findall(r'(?m)^\s*(\d+)\b', block)]
                if idxs:
                    return max(idxs)
        
        # 2) Fallback: search the region just before Numeric Usage Summary
        m_num = numeric_usage_hdr.search(text)
        if not m_num:
            return None
        
        start_idx = max(0, m_num.start() - 120000)
        region = text[start_idx:m_num.start()]
        
        idxs = [int(m.group(1)) for m in bit_use_row.finditer(region)]
        if idxs:
            return max(idxs)
        
        return None
    
    @staticmethod
    def extract_numeric_bits_count(text):
        """Row count from *** Numeric Usage Summary"""
        m = re.search(r'(?is)\*{3}\s*Numeric\s+Usage\s+Summary.*?(?=\n\s*\*{3}|\Z)', text)
        if not m:
            return 0
        block = m.group(0)
        idxs = [int(x) for x in re.findall(r'(?m)^\s*(\d+)\b', block)]
        return len(idxs)
    
    @staticmethod
    def extract_timer_last_index(text):
        """In TIMER BITS section, find the largest -N- marker before the next section"""
        # Find start of BOOLEAN BITS section to narrow down the search window
        bool_start = re.search(r'(?mi)^\s*(?:\d+\s+)?BOOLEAN\s+BITS\s*$', text)
        search_start = bool_start.end() if bool_start else 0

        # 1. Find the start of the section after BOOLEAN BITS
        m_start = re.search(r'(?mi)^\s*(?:\d+\s+)?TIMER\s+BITS\s*$', text[search_start:])
        if not m_start:
            return 0

        # 2. Get the text strictly after "TIMER BITS"
        rest = text[search_start + m_start.end():]

        # 3. Find the nearest section terminator
        # Common end markers in MLL files:
        # - LOG BITS, COMMANDS, CONSTANTS (users suggestion)
        # - *** Usage Summary (and similar asterisks headers)
        # - ======= (Separators)
        # - SECTION (sometimes used)
        terminators = [
            r'LOG\s+BITS',
            r'CONSTANTS'
        ]
        pattern = '|'.join(terminators)

        m_end = re.search(pattern, rest, flags=re.IGNORECASE)

        if m_end:
            # Section ends at the start of the match
            block = rest[:m_end.start()]
        else:
            # No clear terminator found; be cautious.
            # We can either assume the rest of the file is the section (risky if page numbers exist),
            # or limit the search window. 
            # For now, let's look at the next 5000 chars as a heuristic limit if no terminator found,
            # or just take the rest if it's small.
            # However, usually there IS a terminator.
            block = rest

        # 4. Find all -N- indices in the isolated block
        nums = [int(n) for n in re.findall(r'-\s*(\d+)\s*-', block)]
        if not nums:
            return 0
            
        return max(nums)
    
    @staticmethod
    def count_numeric_blocks(text):
        """Count BLOCKs inside NUMERIC BEGIN ... END NUMERIC sections"""
        line_num_re_prefix = r'(?:\s*\d+\s+)?'
        
        total = 0
        cursor = 0
        up = text.upper()
        while True:
            start = up.find('NUMERIC BEGIN', cursor)
            if start < 0:
                break
            end = up.find('END NUMERIC', start)
            if end < 0:
                break
            section = text[start:end]
            total += len(re.findall(rf'(?mi)^{line_num_re_prefix}BLOCK\s+\d+\b', section))
            cursor = end + len('END NUMERIC')
        return total
    
    @staticmethod
    def count_assigns(text):
        """Count ASSIGN and NV.ASSIGN lines"""
        line_num_re_prefix = r'(?:\s*\d+\s+)?'
        
        assign_count = len(re.findall(rf'(?m)^{line_num_re_prefix}ASSIGN\b', text))
        nv_assign_count = len(re.findall(rf'(?m)^{line_num_re_prefix}NV\.ASSIGN\b', text))
        return assign_count, nv_assign_count
    
    @staticmethod
    def count_boards(text):
        """Count TYPE: IN16 / OUT16 / NV.IN32.OUT32 occurrences,
        skipping boards whose ADJUSTABLE ENABLE is 0.

        Implementation strategy:
        1) Start from the original global TYPE counts (robust to format).
        2) Find BOARD blocks that have ADJUSTABLE ENABLE: 0 and subtract their
           contribution from the totals.
        """

        line_num_re_prefix = r'(?:\s*\d+\s+)?'

        # 1) Base counts: all TYPE lines in the file (original behavior)
        vi = len(re.findall(rf'(?mi)^{line_num_re_prefix}TYPE\s*:\s*IN16\b', text))
        vo = len(re.findall(rf'(?mi)^{line_num_re_prefix}TYPE\s*:\s*OUT16\b', text))
        nv = len(re.findall(rf'(?mi)^{line_num_re_prefix}TYPE\s*:\s*NV\.IN32\.OUT32\b', text))

        # 2) Look for disabled BOARD blocks and subtract them.
        # Allow optional leading line numbers before 'BOARD:' to match
        # listings like ' 263    BOARD: J3'.
        boards = list(re.finditer(
            r'(?ms)^\s*\d*\s*BOARD\s*:\s*(\S+)\s*(.*?)(?=^\s*\d*\s*BOARD\s*:|\Z)',
            text
        ))

        for m in boards:
            body = m.group(2)

            # Only adjust counts for disabled boards
            if not re.search(r'(?i)ADJUSTABLE\s+ENABLE\s*:\s*0', body):
                continue

            t = re.search(r'(?i)TYPE\s*:\s*([A-Z0-9\.\-]+)', body)
            btype = (t.group(1).upper() if t else '')

            if 'IN16' in btype and vi > 0:
                vi -= 1
            elif 'OUT16' in btype and vo > 0:
                vo -= 1
            elif ('NV.IN32.OUT32' in btype or 'NV' in btype) and nv > 0:
                nv -= 1

        return vi, vo, nv
    
    @staticmethod
    def count_mii_links(text):
        """Count ADJUSTABLE MII.ADDRESS and ADJUSTABLE MII.NV.ADDRESS"""
        line_num_re_prefix = r'(?:\s*\d+\s+)?'
        
        mi = len(re.findall(rf'(?mi)^{line_num_re_prefix}ADJUSTABLE\s+MII\.ADDRESS\s*:', text))
        mnv = len(re.findall(rf'(?mi)^{line_num_re_prefix}ADJUSTABLE\s+MII\.NV\.ADDRESS\s*:', text))
        return mi, mnv
    
    @staticmethod
    def compute_sync_message_size(booleans, numerics, timers, blocks):
        """SYNC Message size = Booleans*2 + Numerics*6 + Timers*12 + Blocks*12"""
        return booleans * 2 + numerics * 6 + timers * 12 + blocks * 12
    
    @classmethod
    def analyze_file(cls, file_path):
        """
        Analyze a single MLL file and return summary data
        
        Returns:
            dict: Analysis results with keys matching HEADERS (snake_case)
        """
        text = cls.read_mll_text(file_path)
        
        # Extract all metrics
        boolean_bits_count = cls.extract_boolean_bits_count(text) or 0
        numeric_bits_count = cls.extract_numeric_bits_count(text)
        numeric_blocks_count = cls.count_numeric_blocks(text)
        timer_last = cls.extract_timer_last_index(text)
        assign_count, nv_assign_count = cls.count_assigns(text)
        vi, vo, nv = cls.count_boards(text)
        mii_vital, mii_nv = cls.count_mii_links(text)
        crc, checksum = cls.extract_crc_checksum(text)
        
        assign_total = assign_count + nv_assign_count
        comm_links_total = mii_vital + mii_nv
        sync_message_size = cls.compute_sync_message_size(
            booleans=boolean_bits_count,
            numerics=numeric_bits_count,
            timers=timer_last,
            blocks=numeric_blocks_count
        )
        
        return {
            'file': Path(file_path).name,
            'boolean_bits_count': boolean_bits_count,
            'numeric_bits_count': numeric_bits_count,
            'numeric_blocks_count': numeric_blocks_count,
            'assign_count': assign_count,
            'nv_assign_count': nv_assign_count,
            'assign_total': assign_total,
            'timer_bits_count': timer_last,
            'vital_input_boards': vi,
            'vital_output_boards': vo,
            'nv_in32_out32_boards': nv,
            'mii_links_count': mii_vital,
            'mii_nv_links_count': mii_nv,
            'comm_links_total': comm_links_total,
            'sync_message_size': sync_message_size,
            'crc': crc or "",
            'checksum': checksum or "",
        }
    
    @classmethod
    def convert_to_export_format(cls, internal_row):
        """Convert internal snake_case keys to ALL-CAPS headers for export"""
        key_to_header = {
            'file': 'FILE',
            'boolean_bits_count': 'BOOLEAN_BITS_COUNT',
            'numeric_bits_count': 'NUMERIC_BITS_COUNT',
            'numeric_blocks_count': 'NUMERIC_BLOCKS_COUNT',
            'assign_count': 'ASSIGN_COUNT',
            'nv_assign_count': 'NV_ASSIGN_COUNT',
            'assign_total': 'ASSIGN_TOTAL',
            'timer_bits_count': 'TIMER_BITS_COUNT',
            'vital_input_boards': 'VITAL_INPUT_BOARDS',
            'vital_output_boards': 'VITAL_OUTPUT_BOARDS',
            'nv_in32_out32_boards': 'NV_IN32_OUT32_BOARDS',
            'mii_links_count': 'MII_LINKS_COUNT',
            'mii_nv_links_count': 'MII_NV_LINKS_COUNT',
            'comm_links_total': 'COMM_LINKS_TOTAL',
            'sync_message_size': 'SYNC_MESSAGE_SIZE',
            'crc': 'CRC',
            'checksum': 'CHECKSUM',
        }
        
        out = {}
        for k, v in internal_row.items():
            hdr = key_to_header.get(k)
            if hdr:
                out[hdr] = v
        
        # Ensure all headers exist (even if blank)
        for hdr in cls.HEADERS:
            out.setdefault(hdr, "")
        return out
        
    def export_to_excel(self, path):
        """Export analysis results to Excel with validation notes and warnings
        
        Args:
            path: Path to save the Excel file
            
        Returns:
            bool: True if export was successful, False otherwise
            
        Raises:
            Exception: If there's an error during export
        """
        try:
            import pandas as pd
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            
            if not hasattr(self, 'mll_rows_export') or not self.mll_rows_export:
                raise ValueError("No data to export. Please parse files first.")
                
            # Create a copy of the data and apply validation
            data = []
            for row in self.mll_rows_export:
                # Convert to dict if it's not already
                row_dict = dict(row) if not isinstance(row, dict) else row
                # Apply validation
                validated_row = self.validate_row(row_dict)
                data.append(validated_row)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Reorder columns to match HEADERS
            columns_order = [col for col in self.HEADERS if col in df.columns]
            df = df[columns_order]
            
            # Create Excel writer
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='MLL Analysis')
                
                # Get workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['MLL Analysis']
                
                # Define styles
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                error_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                warning_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                white_font = Font(color='FFFFFF')
                
                # Style headers
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Apply formatting based on validation
                for row_idx, row in enumerate(data, start=2):  # +2 for 1-based index and header
                    # Check boolean bits validation
                    if 'BOOLEAN_BITS_COUNT' in row and 'Boolean bits' in row.get('VALIDATION_NOTES', ''):
                        cell = worksheet.cell(row=row_idx, column=columns_order.index('BOOLEAN_BITS_COUNT') + 1)
                        cell.fill = error_fill
                        cell.font = white_font
                        
                    # Check for board count warning
                    if 'VITAL_INPUT_BOARDS' in row and '>12 boards' in row.get('VALIDATION_NOTES', ''):
                        cell = worksheet.cell(row=row_idx, column=columns_order.index('VITAL_INPUT_BOARDS') + 1)
                        cell.fill = warning_fill
                        
                        cell = worksheet.cell(row=row_idx, column=columns_order.index('VITAL_OUTPUT_BOARDS') + 1)
                        cell.fill = warning_fill
                        
                        cell = worksheet.cell(row=row_idx, column=columns_order.index('NV_IN32_OUT32_BOARDS') + 1)
                        cell.fill = warning_fill
                    
                    # Check comm links validation
                    if 'COMM_LINKS_TOTAL' in row and 'Communication links' in row.get('VALIDATION_NOTES', ''):
                        cell = worksheet.cell(row=row_idx, column=columns_order.index('COMM_LINKS_TOTAL') + 1)
                        cell.fill = error_fill
                        cell.font = white_font
                
                # Auto-size columns
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    column_letter = get_column_letter(column_cells[0].column)
                    worksheet.column_dimensions[column_letter].width = min(length + 2, 50)  # Cap width at 50
                
                # Freeze header row
                worksheet.freeze_panes = 'A2'
                
                # Add filter
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Set alignment for all cells
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            return True
            
        except ImportError:
            raise Exception("Required packages not found. Please install with: pip install pandas openpyxl")
        except Exception as e:
            raise Exception(f"Error exporting to Excel: {str(e)}")
